#!/usr/bin/env python3
"""Rigenera le liste giocatori in asta.html a partire da un Excel Fantacalcio.

Uso:
  python build_liste.py
  python build_liste.py percorso.xlsx

Se non passi un path, usa l'.xlsx più recente nella cartella dello script.
L'FVM Excel è su 1000 crediti: lo script lo salva così, l'app lo quota
su 300 o 500 e lo arrotonda all'intero.
"""

from __future__ import annotations

import json
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Serve openpyxl: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parent
HTML_PATH = ROOT / "asta.html"
MARKER_START = "/* === PLAYERS_DATA_START === */"
MARKER_END = "/* === PLAYERS_DATA_END === */"

ROLE_CONFIG = {
    "P": {"tiers": None, "label": "Portieri"},
    "D": {"tiers": (9, 8, 20), "label": "Difensori"},  # n_tiers, size, bonus
    "C": {"tiers": (9, 8, 20), "label": "Centrocampisti"},
    "A": {"tiers": (6, 8, 20), "label": "Attaccanti"},
}

# Sempre in fascia bonus, anche sotto la soglia FVM. Se sono già in una fascia
# numerata (es. Calò in C), li sposta in bonus senza rimescolare le altre fasce.
BONUS_EXTRA = {
    "D": ["Doig", "Kossounou", "Ghilardi", "Kamara H."],
    "C": ["Calò", "Akinsanmiro", "Piotrowski", "Unai Gomez"],
}


def find_xlsx(explicit: str | None) -> Path:
    if explicit:
        p = Path(explicit).expanduser().resolve()
        if not p.exists():
            raise FileNotFoundError(f"File non trovato: {p}")
        return p
    candidates = sorted(ROOT.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if not candidates:
        raise FileNotFoundError(f"Nessun .xlsx in {ROOT}")
    return candidates[0]


def load_players(xlsx: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    if "Tutti" not in wb.sheetnames:
        raise ValueError(f"Foglio 'Tutti' mancante. Fogli: {wb.sheetnames}")
    ws = wb["Tutti"]
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if "Id" in vals and "R" in vals and "Nome" in vals and "FVM" in vals:
            header_row = i
            headers = vals
            break
    if header_row is None:
        raise ValueError("Intestazioni Id/R/Nome/Squadra/FVM non trovate")

    col = {h: idx for idx, h in enumerate(headers)}
    needed = ["Id", "R", "Nome", "Squadra", "FVM"]
    for n in needed:
        if n not in col:
            raise ValueError(f"Colonna mancante: {n}")

    players = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        pid = row[col["Id"]]
        if pid is None:
            continue
        role = row[col["R"]]
        if role not in ("P", "D", "C", "A"):
            continue
        nome = row[col["Nome"]]
        squadra = row[col["Squadra"]]
        if not nome or not squadra:
            continue
        fvm = row[col["FVM"]]
        try:
            fvm_val = float(fvm) if fvm is not None else 0.0
        except (TypeError, ValueError):
            fvm_val = 0.0
        players.append(
            {
                "id": int(pid) if not isinstance(pid, str) else pid,
                "r": str(role),
                "nome": str(nome).strip(),
                "squadra": str(squadra).strip(),
                "_fvm": fvm_val,
            }
        )
    return players


def sort_by_fvm(players: list[dict]) -> list[dict]:
    return sorted(players, key=lambda p: (-p["_fvm"], p["nome"].lower()))


def norm_name(s: str) -> str:
    decomposed = unicodedata.normalize("NFKD", s)
    stripped = "".join(c for c in decomposed if not unicodedata.combining(c))
    return " ".join(stripped.lower().split())


def find_by_name(ordered: list[dict], wanted: str) -> dict | None:
    target = norm_name(wanted)
    exact = [p for p in ordered if norm_name(p["nome"]) == target]
    if len(exact) == 1:
        return exact[0]
    if len(exact) > 1:
        print(f"  nome ambiguo (match esatto): {wanted} -> {[p['nome'] for p in exact]}")
        return None
    partial = [
        p
        for p in ordered
        if target in norm_name(p["nome"]) or norm_name(p["nome"]).startswith(target)
    ]
    if len(partial) == 1:
        return partial[0]
    if partial:
        print(f"  nome ambiguo: {wanted} -> {[p['nome'] for p in partial]}")
    else:
        print(f"  non trovato in Excel: {wanted}")
    return None


def apply_bonus_extra(tiers: list[list[dict]], bonus: list[dict], ordered: list[dict], names: list[str]) -> None:
    for wanted in names:
        src = find_by_name(ordered, wanted)
        if src is None:
            continue
        pub = public_player(src)
        if any(p["id"] == pub["id"] for p in bonus):
            continue
        removed = False
        for tier in tiers:
            for i, p in enumerate(tier):
                if p["id"] == pub["id"]:
                    tier.pop(i)
                    removed = True
                    break
            if removed:
                break
        bonus.append(pub)
        where = "spostato da fascia" if removed else "aggiunto sotto soglia"
        print(f"  bonus extra {src['r']}: {pub['nome']} ({pub['squadra']}) — {where}")


def public_player(p: dict) -> dict:
    fvm = p["_fvm"]
    return {
        "id": p["id"],
        "nome": p["nome"],
        "squadra": p["squadra"],
        "fvm": int(round(fvm)),
    }


def build_keepers(players: list[dict]) -> list[dict]:
    keepers = sort_by_fvm([p for p in players if p["r"] == "P"])
    best_by_team: dict[str, dict] = {}
    for p in keepers:
        if p["squadra"] not in best_by_team:
            best_by_team[p["squadra"]] = p
    selected = sort_by_fvm(list(best_by_team.values()))
    return [public_player(p) for p in selected]


def build_tiered(players: list[dict], role: str, n_tiers: int, tier_size: int, bonus: int) -> dict:
    ordered = sort_by_fvm([p for p in players if p["r"] == role])
    need = n_tiers * tier_size + bonus
    sliced = ordered[:need]
    tiers = []
    for i in range(n_tiers):
        chunk = sliced[i * tier_size : (i + 1) * tier_size]
        tiers.append([public_player(p) for p in chunk])
    bonus_list = [public_player(p) for p in sliced[n_tiers * tier_size : need]]
    apply_bonus_extra(tiers, bonus_list, ordered, BONUS_EXTRA.get(role, []))
    return {"tiers": tiers, "bonus": bonus_list}


def build_payload(players: list[dict], xlsx: Path) -> dict:
    d_cfg = ROLE_CONFIG["D"]["tiers"]
    c_cfg = ROLE_CONFIG["C"]["tiers"]
    a_cfg = ROLE_CONFIG["A"]["tiers"]
    stamp = f"{xlsx.name} | {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    return {
        "dataStamp": stamp,
        "P": build_keepers(players),
        "D": build_tiered(players, "D", *d_cfg),
        "C": build_tiered(players, "C", *c_cfg),
        "A": build_tiered(players, "A", *a_cfg),
    }


def format_players_block(payload: dict) -> str:
    # Compact but readable JSON inside the markers — fvm is the 1000-credit quote
    json_str = json.dumps(payload, ensure_ascii=False, indent=2)
    return (
        f"{MARKER_START}\n"
        f"const PLAYERS = {json_str};\n"
        f"{MARKER_END}"
    )


def inject_into_html(block: str) -> None:
    if not HTML_PATH.exists():
        raise FileNotFoundError(f"Manca {HTML_PATH.name}: crea prima asta.html")
    text = HTML_PATH.read_text(encoding="utf-8")
    if MARKER_START not in text or MARKER_END not in text:
        raise ValueError("Marker PLAYERS_DATA_START/END non trovati in asta.html")
    before = text.split(MARKER_START)[0]
    after = text.split(MARKER_END, 1)[1]
    HTML_PATH.write_text(before + block + after, encoding="utf-8")


def main() -> int:
    explicit = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx = find_xlsx(explicit)
    print(f"Lettura: {xlsx.name}")
    players = load_players(xlsx)
    print(f"Giocatori letti: {len(players)}")
    payload = build_payload(players, xlsx)
    print(
        f"P={len(payload['P'])}  "
        f"D={sum(len(t) for t in payload['D']['tiers'])}+{len(payload['D']['bonus'])}  "
        f"C={sum(len(t) for t in payload['C']['tiers'])}+{len(payload['C']['bonus'])}  "
        f"A={sum(len(t) for t in payload['A']['tiers'])}+{len(payload['A']['bonus'])}"
    )
    block = format_players_block(payload)
    if not HTML_PATH.exists():
        # Allow generating the JSON block alone for bootstrap
        print("asta.html assente: stampo solo il blocco dati (nessuna inject)")
        print(block[:200], "...")
        return 0
    inject_into_html(block)
    print(f"Aggiornato: {HTML_PATH.name}")
    print(f"Stamp: {payload['dataStamp']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
